import math
import os
import tempfile
import unittest

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from leeger.excel.ExcelManager import ExcelManager
from leeger.model.league.League import League
from leeger.model.league.Matchup import Matchup
from leeger.model.league.Team import Team
from leeger.model.league.Week import Week
from leeger.model.league.Year import Year
from test.helper.prototypes import getNDefaultOwnersAndTeams


class TestExcelManager(unittest.TestCase):

    def test_league_toExcel_happyPath(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=2002, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with tempfile.TemporaryDirectory() as tempDir:
            fullPath = os.path.join(tempDir, "tmp.xlsx")
            excelManager = ExcelManager(league=league)
            excelManager.leagueToExcel(fullPath)

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)

            self.assertEqual(4, len(workbook.sheetnames))
            self.assertEqual("2000", workbook.sheetnames[0])
            self.assertEqual("2001", workbook.sheetnames[1])
            self.assertEqual("2002", workbook.sheetnames[2])
            self.assertEqual("All Time", workbook.sheetnames[3])

            # open created Excel file to check that it was saved correctly
            workbook = load_workbook(filename=fullPath)
            worksheet = workbook.active

            self.assertEqual(4, len(workbook.sheetnames))
            self.assertEqual("Team Names", worksheet["A1"].value)
            self.assertEqual("a", worksheet["A2"].value)
            self.assertEqual("b", worksheet["A3"].value)

            for year in league.years:
                statsWithTitles = year.statSheet().preferredOrderWithTitle()
                for row, teamId in enumerate([team.id for team in year.teams]):
                    for col, statWithTitle in enumerate(statsWithTitles):
                        char = get_column_letter(col + 2)
                        if row == 1:
                            # check stat header
                            self.assertEqual(statWithTitle[0], worksheet[f"{char}{row}"].value)
                        # check stat value
                        # due to Excel rounding values, we assert that the values are very, very close
                        assert math.isclose(float(statWithTitle[1][teamId]), worksheet[f"{char}{row + 2}"].value,
                                            rel_tol=0.000000000000001)

    def test_league_toExcel_fileAlreadyExists_raisesException(self):
        owners, teams1 = getNDefaultOwnersAndTeams(2)
        teams1[0].name = "a"
        teams1[1].name = "b"
        matchup1 = Matchup(teamAId=teams1[0].id, teamBId=teams1[1].id, teamAScore=1, teamBScore=2)
        week1 = Week(weekNumber=1, matchups=[matchup1])
        year1 = Year(yearNumber=2000, teams=teams1, weeks=[week1])

        teams2 = [Team(ownerId=owners[0].id, name="a2"), Team(ownerId=owners[1].id, name="b2")]
        matchup2 = Matchup(teamAId=teams2[0].id, teamBId=teams2[1].id, teamAScore=1, teamBScore=2)
        week2 = Week(weekNumber=1, matchups=[matchup2])
        year2 = Year(yearNumber=2001, teams=teams2, weeks=[week2])

        teams3 = [Team(ownerId=owners[0].id, name="a3"), Team(ownerId=owners[1].id, name="b3")]
        matchup3 = Matchup(teamAId=teams3[0].id, teamBId=teams3[1].id, teamAScore=1, teamBScore=2)
        week3 = Week(weekNumber=1, matchups=[matchup3])
        year3 = Year(yearNumber=1999, teams=teams3, weeks=[week3])

        league = League(name="", owners=owners, years=[year1, year2, year3])

        with self.assertRaises(FileExistsError) as context:
            with tempfile.TemporaryDirectory() as tempDir:
                fullPath = os.path.join(tempDir, "tmp.xlsx")
                # create a file with this same name/path first
                with open(fullPath, "w") as f:
                    ...
                excelManager = ExcelManager(league=league)
                excelManager.leagueToExcel(fullPath)
        self.assertEqual(f"Cannot create file at path: '{fullPath}' because there is already a file there.",
                         str(context.exception))